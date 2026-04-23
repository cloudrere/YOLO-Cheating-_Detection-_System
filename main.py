#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
作弊监控系统 (Cheating Monitor System)
=====================================
基于 eye_movement / head_pose / mobile_detection 三个检测模块,
提供图片、视频、摄像头三种检测模式的桌面 GUI 应用。

功能:
    - 图片检测 / 视频检测 / 摄像头实时检测
    - 开始、暂停/继续、停止
    - 警报系统: 头部/瞳孔方向偏离超过阈值时间时报警
    - 结果保存: 违规截图 + Excel 报表 (事件明细 + 统计)
    - 检测前 / 检测后双画面
    - 操作日志 + 事件明细表 + 实时统计
依赖:
    pip install PyQt5 opencv-python openpyxl numpy
"""

import sys
import os
import time
from datetime import datetime

import cv2
import numpy as np

from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QLabel, QFileDialog, QTextEdit, QGroupBox, QGridLayout,
    QMessageBox, QStatusBar, QFrame, QDoubleSpinBox, QSpinBox, QFormLayout,
    QTableWidget, QTableWidgetItem, QHeaderView
)
from PyQt5.QtCore import Qt, QThread, pyqtSignal, QMutex
from PyQt5.QtGui import QImage, QPixmap, QFont

try:
    import openpyxl
    from openpyxl.styles import Font as XLFont, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# === 用户现有的三个检测模块 ===
from eye_movement import process_eye_movement
from head_pose import process_head_pose
from mobile_detection import process_mobile_detection


# =========================================================
#                    预览线程 (选择源后立即显示原始画面)
# =========================================================
class PreviewThread(QThread):
    """
    在点击"开始检测"前显示原始画面:
        - 图片: 立即显示一次
        - 视频: 显示第一帧作为预览
        - 摄像头: 持续推送原始帧, 直到 stop()
    预览线程不调用任何检测模块.
    """
    frame_ready     = pyqtSignal(np.ndarray)
    finished_signal = pyqtSignal()

    def __init__(self, source_type, source, parent=None):
        super().__init__(parent)
        self.source_type = source_type
        self.source      = source
        self._running    = True
        self._mutex      = QMutex()

    def stop(self):
        self._mutex.lock(); self._running = False; self._mutex.unlock()

    def _alive(self):
        self._mutex.lock(); r = self._running; self._mutex.unlock(); return r

    def run(self):
        try:
            if self.source_type == 'image':
                frame = cv2.imread(self.source)
                if frame is not None:
                    self.frame_ready.emit(frame)

            elif self.source_type == 'video':
                cap = cv2.VideoCapture(self.source)
                if cap.isOpened():
                    ret, frame = cap.read()
                    if ret:
                        self.frame_ready.emit(frame)
                    cap.release()

            elif self.source_type == 'camera':
                cap = cv2.VideoCapture(int(self.source))
                if cap.isOpened():
                    while self._alive():
                        ret, frame = cap.read()
                        if not ret:
                            break
                        self.frame_ready.emit(frame)
                        self.msleep(30)
                    cap.release()
        except Exception:
            pass
        self.finished_signal.emit()


# =========================================================
#                    检测工作线程
# =========================================================
class DetectionThread(QThread):
    """
    在独立线程中处理图像/视频/摄像头, 避免阻塞 UI.
    通过信号把 (原始帧, 检测后帧, 状态, 警报) 发送到主线程.
    """
    frame_ready     = pyqtSignal(np.ndarray, np.ndarray)   # (before, after)
    status_update   = pyqtSignal(str, str, bool)           # (head, eye, mobile)
    cheat_detected  = pyqtSignal(str, str, np.ndarray)     # (type, detail, snapshot)
    log_message     = pyqtSignal(str, str)                 # (level, text)
    finished_signal = pyqtSignal()

    def __init__(self, source_type, source,
                 head_threshold=3.0, eye_threshold=3.0, mobile_threshold=3.0,
                 frame_skip=0, parent=None):
        super().__init__(parent)
        self.source_type       = source_type        # 'image' | 'video' | 'camera'
        self.source            = source
        self.head_threshold    = float(head_threshold)
        self.eye_threshold     = float(eye_threshold)
        self.mobile_threshold  = float(mobile_threshold)
        self.frame_skip        = max(0, int(frame_skip))   # 仅视频模式生效
        self._running          = True
        self._paused           = False
        self._mutex            = QMutex()

    # --- 线程控制 ---
    def pause(self):
        self._mutex.lock(); self._paused = True;  self._mutex.unlock()

    def resume(self):
        self._mutex.lock(); self._paused = False; self._mutex.unlock()

    def stop(self):
        self._mutex.lock(); self._running = False; self._paused = False; self._mutex.unlock()

    def is_paused(self):
        self._mutex.lock(); p = self._paused; self._mutex.unlock(); return p

    # --- 主循环分派 ---
    def run(self):
        try:
            if self.source_type == 'image':
                self._run_image()
            else:
                self._run_stream()
        except Exception as e:
            self.log_message.emit("ERROR", f"检测线程异常: {e}")
        self.finished_signal.emit()

    # --- 单张图片 ---
    def _run_image(self):
        frame = cv2.imread(self.source)
        if frame is None:
            self.log_message.emit("ERROR", f"无法读取图片: {self.source}")
            return
        original = frame.copy()

        # 图片模式: 先用 None 调用一次获取 angles 做"自校准",
        # 然后再调用一次获取方向字符串.
        self_calib = None
        try:
            _, angles = process_head_pose(frame.copy(), None)
            self_calib = angles
        except Exception:
            pass

        processed, gaze, head, mobile = self._detect_frame(frame, self_calib, use_calib=True)
        self.status_update.emit(head, gaze, mobile)
        self.frame_ready.emit(original, processed)

        # 图片模式一张即检测一次,发现违规直接记录
        if head != "Looking at Screen":
            self.cheat_detected.emit("head",   f"头部: {head}",  processed.copy())
        if gaze != "Looking at Screen":
            self.cheat_detected.emit("eye",    f"瞳孔: {gaze}",  processed.copy())
        if mobile:
            self.cheat_detected.emit("mobile", "检测到手机",       processed.copy())
        self.log_message.emit("INFO", "图片检测完成")

    # --- 视频 / 摄像头 ---
    def _run_stream(self):
        src = self.source if self.source_type == 'video' else int(self.source)
        cap = cv2.VideoCapture(src)
        if not cap.isOpened():
            self.log_message.emit("ERROR", f"无法打开源: {self.source}")
            return

        calibrated_angles  = None
        calibration_frames = 0
        calibration_target = 30 if self.source_type == 'camera' else 0
        video_calibrated   = False   # 视频模式首帧自校准标志
        skip_counter       = 0       # 视频跳帧计数器
        head_bad_t = eye_bad_t = mobile_bad_t = None

        while self._running:
            if self.is_paused():
                self.msleep(80); continue

            ret, frame = cap.read()
            if not ret:
                if self.source_type == 'video':
                    self.log_message.emit("INFO", "视频播放结束")
                break

            # ---- 视频跳帧: 快进读取, 不做检测 ----
            if (self.source_type == 'video'
                    and self.frame_skip > 0
                    and skip_counter > 0):
                skip_counter -= 1
                continue

            original = frame.copy()

            # ---- 视频模式: 第一帧静默自校准 ----
            if self.source_type == 'video' and not video_calibrated:
                try:
                    _, angles = process_head_pose(frame.copy(), None)
                    if angles is not None:
                        calibrated_angles = angles
                except Exception:
                    pass
                video_calibrated = True

            # ---- 摄像头校准阶段 ----
            if self.source_type == 'camera' and calibration_frames < calibration_target:
                cv2.putText(frame, "Calibrating... keep head straight",
                            (30, 40), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 255), 2)
                try:
                    _, angles = process_head_pose(frame, None)
                    if angles is not None:
                        calibrated_angles = angles
                except Exception:
                    pass
                calibration_frames += 1
                if calibration_frames == calibration_target:
                    self.log_message.emit("INFO", "头部姿态校准完成")
                self.frame_ready.emit(original, frame)
                self.msleep(15)
                continue

            processed, gaze, head, mobile = self._detect_frame(
                frame, calibrated_angles, use_calib=True
            )
            self.status_update.emit(head, gaze, mobile)

            now = time.time()

            # ---- 头部超时警报 ----
            if head != "Looking at Screen":
                if head_bad_t is None:
                    head_bad_t = now
                elif now - head_bad_t >= self.head_threshold:
                    self.cheat_detected.emit("head", f"头部偏离屏幕 ({head})", processed.copy())
                    head_bad_t = None
            else:
                head_bad_t = None

            # ---- 眼部超时警报 ----
            if gaze != "Looking at Screen":
                if eye_bad_t is None:
                    eye_bad_t = now
                elif now - eye_bad_t >= self.eye_threshold:
                    self.cheat_detected.emit("eye", f"瞳孔偏离屏幕 ({gaze})", processed.copy())
                    eye_bad_t = None
            else:
                eye_bad_t = None

            # ---- 手机超时警报 ----
            if mobile:
                if mobile_bad_t is None:
                    mobile_bad_t = now
                elif now - mobile_bad_t >= self.mobile_threshold:
                    self.cheat_detected.emit("mobile", "检测到疑似手机使用", processed.copy())
                    mobile_bad_t = None
            else:
                mobile_bad_t = None

            self.frame_ready.emit(original, processed)

            # ---- 视频模式: 本次已检测完, 重置跳帧计数 ----
            if self.source_type == 'video' and self.frame_skip > 0:
                skip_counter = self.frame_skip

            self.msleep(15)

        cap.release()

    # --- 返回值归一化 (防止底层模块返回 tuple 等意外类型) ---
    @staticmethod
    def _as_str(val, default="Looking at Screen"):
        if isinstance(val, str):
            return val
        if isinstance(val, (tuple, list)):
            for x in val:
                if isinstance(x, str):
                    return x
            return default
        if val is None:
            return default
        return str(val)

    @staticmethod
    def _as_bool(val):
        if isinstance(val, bool):
            return val
        if isinstance(val, (tuple, list)):
            for x in val:
                if isinstance(x, bool):
                    return x
            return False
        return bool(val) if val is not None else False

    # --- 单帧三种检测 ---
    def _detect_frame(self, frame, calibrated_angles, use_calib):
        gaze, head, mobile = "Looking at Screen", "Looking at Screen", False

        try:
            frame, raw = process_eye_movement(frame)
            gaze = self._as_str(raw)
        except Exception as e:
            self.log_message.emit("WARN", f"眼动检测异常: {e}")

        try:
            if use_calib:
                frame, raw = process_head_pose(frame, calibrated_angles)
            else:
                frame, raw = process_head_pose(frame, None)
            head = self._as_str(raw)
        except Exception as e:
            self.log_message.emit("WARN", f"头部姿态检测异常: {e}")

        try:
            frame, raw = process_mobile_detection(frame)
            mobile = self._as_bool(raw)
        except Exception as e:
            self.log_message.emit("WARN", f"手机检测异常: {e}")

        # 在画面上叠加状态信息
        cv2.putText(frame, f"Gaze:   {gaze}",    (20, 30),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.65, (0, 255, 0), 2)
        cv2.putText(frame, f"Head:   {head}",    (20, 60),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.65, (0, 255, 0), 2)
        cv2.putText(frame, f"Mobile: {mobile}",  (20, 90),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.65,
                    (0, 0, 255) if mobile else (0, 255, 0), 2)
        return frame, gaze, head, mobile


# =========================================================
#                    主窗口
# =========================================================
class MainWindow(QMainWindow):

    def __init__(self):
        super().__init__()
        self.setWindowTitle("作弊监控系统 - Cheating Monitor System")
        self.resize(1440, 880)

        self.thread          = None
        self.preview_thread  = None
        self.session_dir     = None
        self.session_records = []     # [{time, type, detail, file}]
        self.counts          = {"head": 0, "eye": 0, "mobile": 0}
        self.current_mode    = None
        self.current_source  = None

        self._build_ui()
        self._apply_theme()
        self._log("INFO", "系统启动完成,请选择检测来源")

    # ---------- UI 构建 ----------
    def _build_ui(self):
        central = QWidget(); self.setCentralWidget(central)
        root = QHBoxLayout(central)
        root.addWidget(self._build_control_panel(), 0)
        root.addWidget(self._build_video_panel(),   3)
        root.addWidget(self._build_right_panel(),   1)
        self.status = QStatusBar(); self.setStatusBar(self.status)
        self.status.showMessage("就绪")

    def _build_control_panel(self):
        box = QGroupBox("控制面板")
        box.setFixedWidth(240)
        v = QVBoxLayout(box)

        # 检测来源
        g_src = QGroupBox("检测来源")
        l1 = QVBoxLayout(g_src)
        self.btn_image  = QPushButton("① 图片检测")
        self.btn_video  = QPushButton("② 视频检测")
        self.btn_camera = QPushButton("③ 摄像头检测")
        self.btn_image .clicked.connect(self._pick_image)
        self.btn_video .clicked.connect(self._pick_video)
        self.btn_camera.clicked.connect(self._pick_camera)
        for b in (self.btn_image, self.btn_video, self.btn_camera):
            l1.addWidget(b)
        v.addWidget(g_src)

        # 检测控制
        g_run = QGroupBox("检测控制")
        l2 = QVBoxLayout(g_run)
        self.btn_start = QPushButton("▶ 开始检测")
        self.btn_pause = QPushButton("⏸ 暂停检测")
        self.btn_stop  = QPushButton("■ 停止检测")
        self.btn_start.clicked.connect(self._start_detection)
        self.btn_pause.clicked.connect(self._toggle_pause)
        self.btn_stop .clicked.connect(self._stop_detection)
        for b in (self.btn_start, self.btn_pause, self.btn_stop):
            b.setEnabled(False); l2.addWidget(b)
        v.addWidget(g_run)

        # 参数设置
        g_set = QGroupBox("参数设置")
        f = QFormLayout(g_set)

        self.spin_head = QDoubleSpinBox()
        self.spin_head.setRange(0.5, 30.0); self.spin_head.setSingleStep(0.5)
        self.spin_head.setValue(3.0);       self.spin_head.setSuffix(" 秒")
        f.addRow("头部阈值:", self.spin_head)

        self.spin_eye = QDoubleSpinBox()
        self.spin_eye.setRange(0.5, 30.0);  self.spin_eye.setSingleStep(0.5)
        self.spin_eye.setValue(3.0);        self.spin_eye.setSuffix(" 秒")
        f.addRow("瞳孔阈值:", self.spin_eye)

        self.spin_mobile = QDoubleSpinBox()
        self.spin_mobile.setRange(0.5, 30.0); self.spin_mobile.setSingleStep(0.5)
        self.spin_mobile.setValue(3.0);       self.spin_mobile.setSuffix(" 秒")
        f.addRow("手机阈值:", self.spin_mobile)

        self.spin_skip = QSpinBox()
        self.spin_skip.setRange(0, 30);  self.spin_skip.setValue(2)
        self.spin_skip.setSuffix(" 帧")
        self.spin_skip.setToolTip("视频模式: 每检测 1 帧, 跳过 N 帧.\n0=不跳帧(逐帧), 2=检1跳2(3x加速)")
        f.addRow("视频跳帧:", self.spin_skip)

        v.addWidget(g_set)

        # 结果输出
        g_save = QGroupBox("结果输出")
        l3 = QVBoxLayout(g_save)
        self.btn_save  = QPushButton("💾 保存结果 (Excel)")
        self.btn_clear = QPushButton("🗑 清除记录")
        self.btn_save .clicked.connect(self._save_results)
        self.btn_clear.clicked.connect(self._clear_records)
        l3.addWidget(self.btn_save); l3.addWidget(self.btn_clear)
        v.addWidget(g_save)

        v.addStretch()

        for b in (self.btn_image, self.btn_video, self.btn_camera,
                  self.btn_start, self.btn_pause, self.btn_stop,
                  self.btn_save,  self.btn_clear):
            b.setMinimumHeight(34)

        self.btn_start.setStyleSheet("background:#27ae60;color:white;font-weight:bold;")
        self.btn_pause.setStyleSheet("background:#f39c12;color:white;font-weight:bold;")
        self.btn_stop .setStyleSheet("background:#c0392b;color:white;font-weight:bold;")
        self.btn_save .setStyleSheet("background:#2980b9;color:white;font-weight:bold;")
        return box

    def _build_video_panel(self):
        box = QGroupBox("检测画面")
        v = QVBoxLayout(box)

        row = QHBoxLayout()
        self.pnl_before = self._make_video_label("检测前 (原始画面)")
        self.pnl_after  = self._make_video_label("检测后 (标注画面)")
        row.addWidget(self.pnl_before)
        row.addWidget(self.pnl_after)
        v.addLayout(row)

        # 三个状态指示灯
        alarm = QGroupBox("实时状态")
        ag = QHBoxLayout(alarm)
        self.lbl_head_st   = QLabel("头部: --")
        self.lbl_eye_st    = QLabel("瞳孔: --")
        self.lbl_mobile_st = QLabel("手机: --")
        for l in (self.lbl_head_st, self.lbl_eye_st, self.lbl_mobile_st):
            l.setAlignment(Qt.AlignCenter)
            l.setFont(QFont("Microsoft YaHei", 11, QFont.Bold))
            l.setMinimumHeight(38)
            ag.addWidget(l)
        self._set_status_ok(self.lbl_head_st,   True)
        self._set_status_ok(self.lbl_eye_st,    True)
        self._set_status_ok(self.lbl_mobile_st, True)
        v.addWidget(alarm)

        return box

    def _make_video_label(self, title):
        panel = QFrame(); panel.setFrameShape(QFrame.StyledPanel)
        l = QVBoxLayout(panel)
        cap = QLabel(title)
        cap.setAlignment(Qt.AlignCenter)
        cap.setFont(QFont("Microsoft YaHei", 10, QFont.Bold))
        img = QLabel("无图像")
        img.setAlignment(Qt.AlignCenter)
        img.setMinimumSize(480, 360)
        img.setStyleSheet("background:#1e1e1e;color:#888;border:1px solid #333;")
        l.addWidget(cap); l.addWidget(img, 1)
        panel.img = img
        return panel

    def _build_right_panel(self):
        w = QWidget(); v = QVBoxLayout(w)

        # 统计卡片
        stat = QGroupBox("结果统计")
        g = QGridLayout(stat)
        self.lbl_cnt_head   = self._counter("头部违规",  "#e67e22")
        self.lbl_cnt_eye    = self._counter("瞳孔违规",  "#8e44ad")
        self.lbl_cnt_mobile = self._counter("手机违规",  "#c0392b")
        self.lbl_cnt_total  = self._counter("总违规数",  "#2c3e50")
        g.addWidget(self.lbl_cnt_head,   0, 0)
        g.addWidget(self.lbl_cnt_eye,    0, 1)
        g.addWidget(self.lbl_cnt_mobile, 1, 0)
        g.addWidget(self.lbl_cnt_total,  1, 1)
        v.addWidget(stat)

        # 事件明细
        evt = QGroupBox("事件明细")
        lg = QVBoxLayout(evt)
        self.tbl_events = QTableWidget(0, 3)
        self.tbl_events.setHorizontalHeaderLabels(["时间", "类型", "说明"])
        self.tbl_events.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.tbl_events.verticalHeader().setVisible(False)
        self.tbl_events.setEditTriggers(QTableWidget.NoEditTriggers)
        lg.addWidget(self.tbl_events)
        v.addWidget(evt, 1)

        # 操作日志
        logb = QGroupBox("操作日志")
        ll = QVBoxLayout(logb)
        self.txt_log = QTextEdit()
        self.txt_log.setReadOnly(True)
        self.txt_log.setFont(QFont("Consolas", 9))
        self.txt_log.setStyleSheet("background:#0f1115;color:#d0d0d0;")
        ll.addWidget(self.txt_log)
        v.addWidget(logb, 1)
        return w

    def _counter(self, title, color):
        lbl = QLabel(f"{title}\n0")
        lbl.setAlignment(Qt.AlignCenter)
        lbl.setFont(QFont("Microsoft YaHei", 12, QFont.Bold))
        lbl.setMinimumHeight(66)
        lbl.setStyleSheet(f"background:{color};color:white;border-radius:6px;padding:6px;")
        return lbl

    def _apply_theme(self):
        self.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                border: 1px solid #c0c0c0;
                border-radius: 5px;
                margin-top: 10px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px; padding: 0 5px;
            }
            QPushButton { border-radius: 4px; }
            QPushButton:disabled { background:#bdbdbd; color:#fff; }
        """)

    # ---------- 来源选择 ----------
    def _pick_image(self):
        fn, _ = QFileDialog.getOpenFileName(self, "选择图片", "", "图片 (*.jpg *.jpeg *.png *.bmp)")
        if fn:
            self.current_mode = 'image'; self.current_source = fn
            self.btn_start.setEnabled(True)
            self.status.showMessage(f"已选择图片: {fn}")
            self._log("INFO", f"选择图片: {os.path.basename(fn)}")
            self._start_preview()

    def _pick_video(self):
        fn, _ = QFileDialog.getOpenFileName(self, "选择视频", "", "视频 (*.mp4 *.avi *.mov *.mkv)")
        if fn:
            self.current_mode = 'video'; self.current_source = fn
            self.btn_start.setEnabled(True)
            self.status.showMessage(f"已选择视频: {fn}")
            self._log("INFO", f"选择视频: {os.path.basename(fn)}")
            self._start_preview()

    def _pick_camera(self):
        self.current_mode = 'camera'; self.current_source = 0
        self.btn_start.setEnabled(True)
        self.status.showMessage("已选择摄像头 (index=0)")
        self._log("INFO", "选择摄像头作为输入源")
        self._start_preview()

    # ---------- 预览控制 ----------
    def _start_preview(self):
        """选择源后启动预览, 只显示原始画面到"检测前"窗口."""
        self._stop_preview()
        self.pnl_after.img.clear()
        self.pnl_after.img.setText("等待开始检测...")
        self.preview_thread = PreviewThread(self.current_mode, self.current_source)
        self.preview_thread.frame_ready.connect(self._on_preview_frame)
        self.preview_thread.start()

    def _stop_preview(self):
        if self.preview_thread and self.preview_thread.isRunning():
            self.preview_thread.stop()
            self.preview_thread.wait(1000)
        self.preview_thread = None

    def _on_preview_frame(self, frame):
        # 预览只更新"检测前"窗口
        self._show(self.pnl_before.img, frame)

    # ---------- 检测控制 ----------
    def _start_detection(self):
        if self.current_mode is None:
            QMessageBox.warning(self, "提示", "请先选择检测来源"); return
        if self.thread and self.thread.isRunning(): return

        # 停止预览, 交由检测线程接管
        self._stop_preview()

        self._new_session()
        self.thread = DetectionThread(
            self.current_mode, self.current_source,
            head_threshold   = self.spin_head.value(),
            eye_threshold    = self.spin_eye.value(),
            mobile_threshold = self.spin_mobile.value(),
            frame_skip       = self.spin_skip.value(),
        )
        self.thread.frame_ready     .connect(self._on_frame)
        self.thread.status_update   .connect(self._on_status)
        self.thread.cheat_detected  .connect(self._on_cheat)
        self.thread.log_message     .connect(self._log)
        self.thread.finished_signal .connect(self._on_thread_finished)
        self.thread.start()

        self.btn_start .setEnabled(False)
        self.btn_pause .setEnabled(self.current_mode != 'image')
        self.btn_stop  .setEnabled(True)
        self.btn_image .setEnabled(False)
        self.btn_video .setEnabled(False)
        self.btn_camera.setEnabled(False)
        self.btn_pause .setText("⏸ 暂停检测")
        self._log("INFO",
                  f"开始检测 (模式={self.current_mode}, "
                  f"头={self.spin_head.value()}s 眼={self.spin_eye.value()}s "
                  f"手机={self.spin_mobile.value()}s 跳帧={self.spin_skip.value()})")

    def _toggle_pause(self):
        if not self.thread: return
        if self.thread.is_paused():
            self.thread.resume()
            self.btn_pause.setText("⏸ 暂停检测")
            self._log("INFO", "继续检测")
        else:
            self.thread.pause()
            self.btn_pause.setText("▶ 继续检测")
            self._log("INFO", "暂停检测")

    def _stop_detection(self):
        if self.thread:
            self.thread.stop()
            self._log("INFO", "请求停止检测")

    def _on_thread_finished(self):
        self.btn_start .setEnabled(True)
        self.btn_pause .setEnabled(False)
        self.btn_stop  .setEnabled(False)
        self.btn_image .setEnabled(True)
        self.btn_video .setEnabled(True)
        self.btn_camera.setEnabled(True)
        self._log("INFO", "检测已停止")
        self.status.showMessage("检测结束")
        # 检测结束后恢复预览; 图片模式除外, 保持检测后画面可见
        if self.current_mode is not None and self.current_mode != 'image':
            self._start_preview()

    # ---------- 信号回调 ----------
    def _on_frame(self, before, after):
        self._show(self.pnl_before.img, before)
        self._show(self.pnl_after .img, after)

    def _on_status(self, head, eye, mobile):
        self.lbl_head_st  .setText(f"头部: {head}")
        self.lbl_eye_st   .setText(f"瞳孔: {eye}")
        self.lbl_mobile_st.setText(f"手机: {'是' if mobile else '否'}")
        self._set_status_ok(self.lbl_head_st,   head == "Looking at Screen")
        self._set_status_ok(self.lbl_eye_st,    eye  == "Looking at Screen")
        self._set_status_ok(self.lbl_mobile_st, not mobile)

    def _set_status_ok(self, lbl, ok):
        color = "#27ae60" if ok else "#c0392b"
        lbl.setStyleSheet(f"background:{color};color:white;border-radius:4px;")

    def _on_cheat(self, ctype, detail, frame):
        self.counts[ctype] = self.counts.get(ctype, 0) + 1
        self._refresh_counts()
        ts = datetime.now()
        fname = ""
        if self.session_dir:
            os.makedirs(self.session_dir, exist_ok=True)
            fname = os.path.join(
                self.session_dir,
                f"{ctype}_{ts.strftime('%Y%m%d_%H%M%S_%f')}.png"
            )
            try:
                cv2.imwrite(fname, frame)
            except Exception as e:
                self._log("WARN", f"截图保存失败: {e}"); fname = ""
        rec = {
            "time":   ts.strftime("%Y-%m-%d %H:%M:%S"),
            "type":   ctype,
            "detail": detail,
            "file":   fname
        }
        self.session_records.append(rec)
        self._add_event_row(rec)
        self._log("ALARM",
                  f"[警报] {detail}  截图={os.path.basename(fname) if fname else '(未保存)'}")

    def _add_event_row(self, rec):
        r = self.tbl_events.rowCount()
        self.tbl_events.insertRow(r)
        self.tbl_events.setItem(r, 0, QTableWidgetItem(rec["time"]))
        self.tbl_events.setItem(r, 1, QTableWidgetItem(rec["type"]))
        self.tbl_events.setItem(r, 2, QTableWidgetItem(rec["detail"]))
        self.tbl_events.scrollToBottom()

    def _refresh_counts(self):
        self.lbl_cnt_head  .setText(f"头部违规\n{self.counts['head']}")
        self.lbl_cnt_eye   .setText(f"瞳孔违规\n{self.counts['eye']}")
        self.lbl_cnt_mobile.setText(f"手机违规\n{self.counts['mobile']}")
        self.lbl_cnt_total .setText(f"总违规数\n{sum(self.counts.values())}")

    # ---------- 保存 / 清除 ----------
    def _new_session(self):
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.session_dir = os.path.join("results", f"session_{ts}")
        os.makedirs(self.session_dir, exist_ok=True)
        self.session_records = []
        self.counts = {"head": 0, "eye": 0, "mobile": 0}
        self._refresh_counts()
        self.tbl_events.setRowCount(0)
        self._log("INFO", f"创建会话目录: {self.session_dir}")

    def _save_results(self):
        if not self.session_records:
            QMessageBox.information(self, "提示", "当前没有可保存的记录"); return
        if not OPENPYXL_AVAILABLE:
            QMessageBox.warning(self, "缺少依赖",
                "未安装 openpyxl,无法导出 Excel.\n请运行: pip install openpyxl")
            return
        default_path = os.path.join(
            self.session_dir or "results",
            f"report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        fn, _ = QFileDialog.getSaveFileName(self, "保存 Excel 报表", default_path, "Excel (*.xlsx)")
        if not fn: return
        try:
            self._write_excel(fn)
            self._log("INFO", f"报表已保存: {fn}")
            QMessageBox.information(self, "成功", f"报表已保存:\n{fn}")
        except Exception as e:
            self._log("ERROR", f"保存失败: {e}")
            QMessageBox.critical(self, "错误", f"保存失败: {e}")

    def _write_excel(self, path):
        wb = openpyxl.Workbook()
        ws1 = wb.active; ws1.title = "事件明细"
        headers = ["序号", "时间", "类型", "说明", "截图文件"]
        ws1.append(headers)
        for c in range(1, len(headers) + 1):
            cell = ws1.cell(row=1, column=c)
            cell.font = XLFont(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="2C3E50")
            cell.alignment = Alignment(horizontal="center")
        for i, r in enumerate(self.session_records, 1):
            ws1.append([i, r["time"], r["type"], r["detail"], r["file"]])
        for col, w in zip("ABCDE", [6, 22, 10, 30, 44]):
            ws1.column_dimensions[col].width = w

        ws2 = wb.create_sheet("结果统计")
        ws2.append(["类型", "次数"])
        ws2.append(["头部违规", self.counts["head"]])
        ws2.append(["瞳孔违规", self.counts["eye"]])
        ws2.append(["手机违规", self.counts["mobile"]])
        ws2.append(["总计",    sum(self.counts.values())])
        for c in range(1, 3):
            ws2.cell(row=1, column=c).font = XLFont(bold=True, color="FFFFFF")
            ws2.cell(row=1, column=c).fill = PatternFill("solid", fgColor="2C3E50")
        ws2.column_dimensions["A"].width = 16
        ws2.column_dimensions["B"].width = 10
        wb.save(path)

    def _clear_records(self):
        if self.thread and self.thread.isRunning():
            QMessageBox.warning(self, "提示", "请先停止检测"); return
        self.counts = {"head": 0, "eye": 0, "mobile": 0}
        self.session_records = []
        self.tbl_events.setRowCount(0)
        self._refresh_counts()
        self._log("INFO", "已清除统计与事件记录")

    # ---------- 日志 / 显示 ----------
    def _log(self, level, msg):
        ts = datetime.now().strftime("%H:%M:%S")
        color = {"INFO": "#7fb6f5", "WARN": "#f5d67f",
                 "ERROR": "#f57f7f", "ALARM": "#ff6b6b"}.get(level, "#d0d0d0")
        self.txt_log.append(
            f'<span style="color:#888">[{ts}]</span> '
            f'<span style="color:{color};font-weight:bold">[{level}]</span> '
            f'<span style="color:#eaeaea">{msg}</span>'
        )

    def _show(self, lbl, frame_bgr):
        if frame_bgr is None: return
        rgb = cv2.cvtColor(frame_bgr, cv2.COLOR_BGR2RGB)
        h, w, ch = rgb.shape
        qimg = QImage(rgb.data, w, h, ch * w, QImage.Format_RGB888)
        pix = QPixmap.fromImage(qimg).scaled(
            lbl.width(), lbl.height(), Qt.KeepAspectRatio, Qt.SmoothTransformation
        )
        lbl.setPixmap(pix)

    # ---------- 关闭 ----------
    def closeEvent(self, event):
        self._stop_preview()
        if self.thread and self.thread.isRunning():
            self.thread.stop()
            self.thread.wait(2000)
        event.accept()


# =========================================================
#                    程序入口
# =========================================================
def main():
    app = QApplication(sys.argv)
    app.setFont(QFont("Microsoft YaHei", 9))
    w = MainWindow()
    w.show()
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()